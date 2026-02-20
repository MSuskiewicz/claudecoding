#!/usr/bin/env python3
"""
simulate_multimer.py

Compute and plot the equilibrium distribution of oligomeric species
(1-mer through N-mer) for a filamentous multimer with a constant
step-wise dissociation constant (Kd) between each protomer addition.

Model
-----
  A1 + A1  <-->  A2    Kd = [A1]^2 / [A2]
  A2 + A1  <-->  A3    Kd = [A2][A1] / [A3]
  ...
  A(n-1) + A1  <-->  An

This gives  [An] = [A1]^n / Kd^(n-1)

Conservation: C_total = sum_{n=1}^{N} n * [An]

[A1] is found by numerical root-finding; all other species follow analytically.

Usage
-----
  python simulate_multimer.py --Kd 1.0 --Ctotal 10.0 --Nmax 12

Arguments
---------
  --Kd      Dissociation constant in µM  (default: 1.0)
  --Ctotal  Total protein concentration in µM  (default: 10.0)
  --Nmax    Largest oligomer to include  (default: 12)
  --out     Output image filename  (default: multimer_distribution.png)
  --yaxis   Y-axis quantity: 'fraction' (fraction of total protein mass)
            or 'conc' (molar concentration in µM)  (default: fraction)
  --excel   Output Excel calculator filename  (default: multimer_calculator.xlsx)
"""

import argparse
import sys
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import brentq

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


# ── Core physics ─────────────────────────────────────────────────────────────

def species_concentrations(m1, Kd, Nmax):
    """
    Return list of concentrations [A1], [A2], ..., [ANmax] in µM.

    [An] = m1^n / Kd^(n-1)
    """
    return [m1 ** n / Kd ** (n - 1) for n in range(1, Nmax + 1)]


def total_monomer_equiv(m1, Kd, Nmax):
    """
    Sum_{n=1}^{Nmax} n * [An]  --  monomer-equivalent total concentration.
    """
    cn = species_concentrations(m1, Kd, Nmax)
    return sum((n + 1) * c for n, c in enumerate(cn))


def solve_monomer(Ctotal, Kd, Nmax):
    """
    Find the free monomer concentration [A1] that satisfies the conservation
    equation by bisection.
    """
    lo, hi = 1e-12, Ctotal
    m1 = brentq(lambda m: total_monomer_equiv(m, Kd, Nmax) - Ctotal,
                lo, hi, xtol=1e-14, rtol=1e-12)
    return m1


def weighted_avg_size(concentrations):
    """
    Return (number-weighted average n, mass-weighted average n).

    Number-weighted: <n> = sum(n * [An]) / sum([An])
    Mass-weighted:  <n>_m = sum(n^2 * [An]) / sum(n * [An])
    """
    n_values = list(range(1, len(concentrations) + 1))
    sum_c   = sum(concentrations)
    sum_nc  = sum(n * c for n, c in zip(n_values, concentrations))
    sum_n2c = sum(n * n * c for n, c in zip(n_values, concentrations))
    n_avg_number = sum_nc  / sum_c   if sum_c  > 0 else 0.0
    n_avg_mass   = sum_n2c / sum_nc  if sum_nc > 0 else 0.0
    return n_avg_number, n_avg_mass


# ── Plotting ─────────────────────────────────────────────────────────────────

def plot_distribution(concentrations, Ctotal, Kd, Nmax, yaxis, outfile):
    n_values = list(range(1, Nmax + 1))

    if yaxis == "fraction":
        monomer_equiv = [n * c for n, c in zip(n_values, concentrations)]
        total = sum(monomer_equiv)
        y = [v / total for v in monomer_equiv]
        ylabel = "Fraction of total protein"
    else:
        y = concentrations
        ylabel = "Concentration (µM)"

    n_avg_num, n_avg_mass = weighted_avg_size(concentrations)

    fig, ax = plt.subplots(figsize=(max(6, Nmax * 0.7), 5))
    bars = ax.bar(n_values, y, color="steelblue", edgecolor="white", linewidth=0.5)

    ax.set_xlabel("Oligomeric state (n-mer)", fontsize=12)
    ax.set_ylabel(ylabel, fontsize=12)
    ax.set_title(
        f"Multimer distribution\n"
        f"Kd = {Kd:.3g} µM   |   C_total = {Ctotal:.3g} µM   |   max n-mer = {Nmax}\n"
        f"<n> (number) = {n_avg_num:.2f}   |   <n> (mass) = {n_avg_mass:.2f}",
        fontsize=10,
    )
    ax.set_xticks(n_values)
    ax.set_xticklabels([f"{n}" for n in n_values])

    threshold = 0.01 if yaxis == "fraction" else 0.001
    for bar, val in zip(bars, y):
        if val >= threshold:
            ax.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height() + max(y) * 0.01,
                f"{val:.2f}" if yaxis == "fraction" else f"{val:.3g}",
                ha="center", va="bottom", fontsize=8,
            )

    ax.set_ylim(0, max(y) * 1.20)
    plt.tight_layout()
    plt.savefig(outfile, dpi=150)
    print(f"Plot saved to {outfile}")
    plt.show()


# ── Excel interactive calculator ─────────────────────────────────────────────

# Fixed table rows for up to MAX_N mers in the Excel sheet
_MAX_N = 100  # supports up to 100-mer in the spreadsheet

# Row indices (1-based, as openpyxl uses)
_R_TITLE   = 1
_R_KD      = 3
_R_CTOTAL  = 4
_R_NMAX    = 5
_R_A1      = 7   # free monomer — solve via Goal Seek
_R_CONS    = 8   # conservation check
_R_INSTR   = 10  # instructions block (10-16)
_R_THEAD   = 18  # table header
_R_DATA0   = 19  # first data row  (n = 1)
# last data row = _R_DATA0 + _MAX_N - 1
_R_TOTAL   = _R_DATA0 + _MAX_N       # totals row
_R_NAVG    = _R_TOTAL + 2            # number-weighted average
_R_MAVG    = _R_TOTAL + 3            # mass-weighted average

# Column indices
_C_A  = 1   # n
_C_B  = 2   # [An]
_C_C  = 3   # n * [An]  (monomer-equiv conc)
_C_D  = 4   # mass fraction
_C_E  = 5   # molar fraction


def _cell(row, col):
    """Return Excel cell address string like 'B7'."""
    return f"{get_column_letter(col)}{row}"


def _abs(row, col):
    """Return absolute Excel cell address like '$B$7'."""
    return f"${get_column_letter(col)}${row}"


def create_excel_calculator(m1, Kd, Ctotal, Nmax, outfile):
    """
    Create a self-contained Excel workbook that:
      - Uses actual Excel formulas for all species concentrations, fractions,
        weighted averages, and the conservation check.
      - Pre-fills the solved [A1] so it opens ready-to-use.
      - Includes Goal Seek instructions so the user can re-solve for new params.
      - Includes an embedded bar chart of mass fractions.
      - Has a 'Model' sheet documenting the isodesmic equations.
    """
    if not _OPENPYXL_AVAILABLE:
        print("openpyxl not installed — skipping Excel export. "
              "Install with: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calculator"

    # ── Style helpers ─────────────────────────────────────────────────────────
    thin    = Side(style="thin",   color="AAAAAA")
    thick   = Side(style="medium", color="2E75B6")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    tborder = Border(left=thick, right=thick, top=thick, bottom=thick)
    center  = Alignment(horizontal="center", vertical="center")
    left    = Alignment(horizontal="left",   vertical="center")

    h_fill  = PatternFill("solid", fgColor="2E75B6")   # dark blue — table header
    h_font  = Font(bold=True, color="FFFFFF", size=11)
    y_fill  = PatternFill("solid", fgColor="FFF2CC")   # yellow — user inputs
    o_fill  = PatternFill("solid", fgColor="FCE4D6")   # orange — goal seek target
    g_fill  = PatternFill("solid", fgColor="E2EFDA")   # green  — computed
    t_fill  = PatternFill("solid", fgColor="BDD7EE")   # totals row
    a_fill  = PatternFill("solid", fgColor="DDEBF7")   # alt row
    i_fill  = PatternFill("solid", fgColor="F2F2F2")   # instructions

    def style(cell, fill=None, font=None, align=center, brd=border):
        if fill:   cell.fill      = fill
        if font:   cell.font      = font
        cell.alignment = align
        cell.border    = brd

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells(f"A{_R_TITLE}:E{_R_TITLE}")
    c = ws[f"A{_R_TITLE}"]
    c.value     = "Isodesmic Multimer Equilibrium — Interactive Calculator"
    c.font      = Font(bold=True, size=14, color="1F3864")
    c.alignment = center
    ws.row_dimensions[_R_TITLE].height = 24

    # ── Parameter input block ─────────────────────────────────────────────────
    params = [
        (_R_KD,    "Kd (µM)",              Kd,    y_fill, "Dissociation constant (µM) — change freely"),
        (_R_CTOTAL,"C_total (µM)",          Ctotal,y_fill, "Total protein concentration (µM) — change freely"),
        (_R_NMAX,  "Max n-mer (Nmax)",      Nmax,  y_fill, "Largest oligomer shown (2 – 20)"),
        (_R_A1,    "Free monomer [A1] (µM)",m1,    o_fill, "← solve via Goal Seek (see instructions)"),
    ]
    for row, label, value, fill, note in params:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=round(value, 8))
        nc = ws.cell(row=row, column=3, value=note)
        lc.font  = Font(bold=True);  lc.fill = fill;  lc.border = border;  lc.alignment = left
        vc.fill  = fill;             vc.border = border; vc.alignment = center
        nc.font  = Font(italic=True, color="595959"); nc.alignment = left

    # ── Conservation check ────────────────────────────────────────────────────
    # B_CONS = Ctotal - SUM(n*[An])  → should be 0
    data_last = _R_DATA0 + _MAX_N - 1
    cons_formula = (f"={_abs(_R_CTOTAL, _C_B)}"
                    f"-SUM({_cell(_R_DATA0, _C_C)}:{_cell(data_last, _C_C)})")
    lc = ws.cell(row=_R_CONS, column=1, value="Conservation check (Goal Seek → 0)")
    vc = ws.cell(row=_R_CONS, column=2, value=cons_formula)
    nc = ws.cell(row=_R_CONS, column=3, value="Use Goal Seek to set this to 0 by changing B7")
    lc.font = Font(bold=True); lc.border = border; lc.alignment = left
    vc.fill = g_fill;          vc.border = tborder; vc.alignment = center
    nc.font = Font(italic=True, color="595959"); nc.alignment = left

    # ── Goal Seek instructions ────────────────────────────────────────────────
    instr_lines = [
        ("HOW TO USE THIS CALCULATOR", True),
        ("1.  Enter Kd, C_total, Nmax in the yellow cells above.", False),
        ("2.  Enter an initial guess for [A1] in B7 (e.g. C_total / 10).", False),
        ("3.  Go to  Data  →  What-If Analysis  →  Goal Seek", False),
        (f"4.  Set cell:  B{_R_CONS}   |   To value:  0   |   By changing cell:  B{_R_A1}", False),
        ("5.  Click OK.  All species concentrations update automatically.", False),
        ("Note: columns B–E in the table are fully formula-driven.", False),
    ]
    for i, (text, bold) in enumerate(instr_lines, start=0):
        r = _R_INSTR + i
        c = ws.cell(row=r, column=1, value=text)
        ws.merge_cells(f"A{r}:E{r}")
        c.font      = Font(bold=bold, size=10 if bold else 9,
                           color="1F3864" if bold else "404040")
        c.fill      = i_fill
        c.alignment = left

    # ── Table header ──────────────────────────────────────────────────────────
    headers = ["n-mer", "[An] (µM)", "n·[An] (µM)", "Mass fraction", "Molar fraction"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=_R_THEAD, column=col, value=h)
        c.font = h_font;  c.fill = h_fill;  c.alignment = center;  c.border = border

    # ── Data rows (n = 1 to _MAX_N) ──────────────────────────────────────────
    # Column refs used in formulas
    kd_ref    = _abs(_R_KD,    _C_B)
    nmax_ref  = _abs(_R_NMAX,  _C_B)
    a1_ref    = _abs(_R_A1,    _C_B)
    sum_B     = f"SUM({_cell(_R_DATA0, _C_B)}:{_cell(data_last, _C_B)})"
    sum_C     = f"SUM({_cell(_R_DATA0, _C_C)}:{_cell(data_last, _C_C)})"
    # mass-weighted denominator is SUM of col C

    for i in range(_MAX_N):
        n   = i + 1
        row = _R_DATA0 + i
        fill = a_fill if i % 2 == 1 else PatternFill()

        # Col A: n
        ca = ws.cell(row=row, column=_C_A, value=n)
        ca.alignment = center;  ca.border = border

        # Col B: [An] = IF(n <= Nmax, A1^n / Kd^(n-1), 0)
        an_ref = _cell(row, _C_A)
        f_B = (f"=IF({an_ref}<={nmax_ref},"
               f"{a1_ref}^{an_ref}/{kd_ref}^({an_ref}-1),0)")
        cb = ws.cell(row=row, column=_C_B, value=f_B)

        # Col C: n * [An]
        f_C = f"={_cell(row, _C_A)}*{_cell(row, _C_B)}"
        cc = ws.cell(row=row, column=_C_C, value=f_C)

        # Col D: mass fraction = n*[An] / total_mass
        f_D = (f"=IF({sum_C}>0,{_cell(row, _C_C)}/{sum_C},0)")
        cd = ws.cell(row=row, column=_C_D, value=f_D)

        # Col E: molar fraction = [An] / total_molar
        f_E = (f"=IF({sum_B}>0,{_cell(row, _C_B)}/{sum_B},0)")
        ce = ws.cell(row=row, column=_C_E, value=f_E)

        for cell in (ca, cb, cc, cd, ce):
            cell.fill = fill;  cell.border = border;  cell.alignment = center

        # Format fractions as percentages
        cd.number_format = "0.00%"
        ce.number_format = "0.00%"
        cb.number_format = "0.000000"
        cc.number_format = "0.000000"

    # ── Totals row ────────────────────────────────────────────────────────────
    tot_labels = ["TOTAL",
                  f"=SUM({_cell(_R_DATA0, _C_B)}:{_cell(data_last, _C_B)})",
                  f"=SUM({_cell(_R_DATA0, _C_C)}:{_cell(data_last, _C_C)})",
                  f"=SUM({_cell(_R_DATA0, _C_D)}:{_cell(data_last, _C_D)})",
                  f"=SUM({_cell(_R_DATA0, _C_E)}:{_cell(data_last, _C_E)})"]
    for col, val in enumerate(tot_labels, start=1):
        c = ws.cell(row=_R_TOTAL, column=col, value=val)
        c.font = Font(bold=True); c.fill = t_fill; c.border = border; c.alignment = center
        if col in (4, 5): c.number_format = "0.00%"
        if col in (2, 3): c.number_format = "0.000000"

    # ── Weighted average sizes ────────────────────────────────────────────────
    # Number-weighted: sum(n*[An]) / sum([An])
    # Mass-weighted:   sum(n^2*[An]) / sum(n*[An])
    n_rng  = f"{_cell(_R_DATA0, _C_A)}:{_cell(data_last, _C_A)}"
    b_rng  = f"{_cell(_R_DATA0, _C_B)}:{_cell(data_last, _C_B)}"
    c_rng  = f"{_cell(_R_DATA0, _C_C)}:{_cell(data_last, _C_C)}"

    avg_rows = [
        (_R_NAVG, "Number-weighted avg size <n>",
         f"=IF({sum_B}>0,SUMPRODUCT({n_rng},{b_rng})/{sum_B},0)",
         "= Σ n·[An] / Σ [An]"),
        (_R_MAVG, "Mass-weighted avg size <n>_mass",
         f"=IF({sum_C}>0,SUMPRODUCT({n_rng},{c_rng})/{sum_C},0)",
         "= Σ n²·[An] / Σ n·[An]"),
    ]
    for row, label, formula, note in avg_rows:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=formula)
        nc = ws.cell(row=row, column=3, value=note)
        lc.font = Font(bold=True); lc.fill = g_fill; lc.border = border; lc.alignment = left
        vc.fill = g_fill;          vc.border = border; vc.alignment = center
        nc.font = Font(italic=True, color="595959"); nc.alignment = left
        vc.number_format = "0.00"

    # ── Bar chart (mass fraction) ─────────────────────────────────────────────
    chart = BarChart()
    chart.type       = "col"
    chart.title      = "Multimer Distribution (mass fraction)"
    chart.y_axis.title = "Mass fraction"
    chart.x_axis.title = "n-mer"
    chart.style      = 10
    chart.width      = 18
    chart.height     = 12

    data_ref = Reference(ws,
                         min_col=_C_D, max_col=_C_D,
                         min_row=_R_THEAD, max_row=data_last)
    cats_ref = Reference(ws,
                         min_col=_C_A, max_col=_C_A,
                         min_row=_R_DATA0, max_row=data_last)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "2E75B6"

    ws.add_chart(chart, f"G{_R_THEAD}")

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {"A": 30, "B": 16, "C": 16, "D": 16, "E": 16,
                  "F": 2,  "G": 5}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Model sheet ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Model")
    ws2.column_dimensions["A"].width = 52
    ws2.column_dimensions["B"].width = 42

    eq_lines = [
        ("Isodesmic (Linear Filament) Model", True, None),
        ("", False, None),
        ("Equilibrium reactions:", True, None),
        ("  A1 + A1 ⇌ A2",        False, "Kd = [A1]² / [A2]"),
        ("  A2 + A1 ⇌ A3",        False, "Kd = [A2][A1] / [A3]"),
        ("  …",                    False, ""),
        ("  A(n-1) + A1 ⇌ An",   False, "Kd = [A(n-1)][A1] / [An]"),
        ("", False, None),
        ("General solution:",      True,  "[An] = [A1]ⁿ / Kd^(n−1)"),
        ("Conservation:",          True,  "C_total = Σ n·[An]  (n = 1…Nmax)"),
        ("", False, None),
        ("Weighted average sizes:", True, None),
        ("  Number-weighted",      False, "<n> = Σ n·[An] / Σ [An]"),
        ("  Mass-weighted",        False, "<n>_mass = Σ n²·[An] / Σ n·[An]"),
        ("", False, None),
        ("[A1] is found by numerical root-finding (Python) or Goal Seek (Excel).", False, None),
        ("All higher species follow analytically from [A1].", False, None),
    ]
    hfont = Font(bold=True, size=12, color="1F3864")
    for r, (a, bold, b) in enumerate(eq_lines, start=1):
        ca = ws2.cell(row=r, column=1, value=a)
        ca.font = hfont if (bold and r == 1) else (Font(bold=bold) if bold else Font())
        if b is not None:
            ws2.cell(row=r, column=2, value=b)

    wb.save(outfile)
    print(f"Excel calculator saved to {outfile}")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Equilibrium distribution of filamentous multimers.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--Kd",     type=float, default=1.0,
                        help="Step-wise dissociation constant in µM (default: 1.0)")
    parser.add_argument("--Ctotal", type=float, default=10.0,
                        help="Total protein concentration in µM (default: 10.0)")
    parser.add_argument("--Nmax",   type=int,   default=12,
                        help="Largest oligomer to include (default: 12)")
    parser.add_argument("--out",    default="multimer_distribution.png",
                        help="Output image file (default: multimer_distribution.png)")
    parser.add_argument("--yaxis",  choices=["fraction", "conc"], default="fraction",
                        help="Y-axis: 'fraction' of total protein or 'conc' in µM (default: fraction)")
    parser.add_argument("--excel",  default="multimer_calculator.xlsx",
                        help="Output Excel calculator file (default: multimer_calculator.xlsx). "
                             "Pass empty string to skip.")
    args = parser.parse_args()

    if args.Kd <= 0:
        sys.exit("--Kd must be positive.")
    if args.Ctotal <= 0:
        sys.exit("--Ctotal must be positive.")
    if args.Nmax < 2:
        sys.exit("--Nmax must be >= 2.")
    if args.Nmax > _MAX_N:
        sys.exit(f"--Nmax must be <= {_MAX_N}.")

    print(f"Parameters: Kd = {args.Kd} µM, C_total = {args.Ctotal} µM, Nmax = {args.Nmax}")

    m1 = solve_monomer(args.Ctotal, args.Kd, args.Nmax)
    concentrations = species_concentrations(m1, args.Kd, args.Nmax)

    # ── Print table ───────────────────────────────────────────────────────────
    total_mass  = sum(n * c for n, c in enumerate(concentrations, 1))
    total_molar = sum(concentrations)
    print(f"\n{'n-mer':>6}  {'[An] (µM)':>14}  {'mass frac':>10}  {'molar frac':>11}")
    print("-" * 50)
    for n, c in enumerate(concentrations, 1):
        mf = n * c / total_mass
        xf = c / total_molar
        print(f"{n:>6}  {c:>14.4g}  {mf:>10.4f}  {xf:>11.4f}")
    print("-" * 50)
    print(f"{'Total':>6}  {total_molar:>14.4g}  {'1.0000':>10}  {'1.0000':>11}")

    # ── Weighted averages ─────────────────────────────────────────────────────
    n_avg_num, n_avg_mass = weighted_avg_size(concentrations)
    print(f"\nNumber-weighted average size: {n_avg_num:.3f}-mer")
    print(f"Mass-weighted   average size: {n_avg_mass:.3f}-mer")

    # ── Plot ──────────────────────────────────────────────────────────────────
    plot_distribution(concentrations, args.Ctotal, args.Kd, args.Nmax,
                      args.yaxis, args.out)

    # ── Excel calculator ──────────────────────────────────────────────────────
    if args.excel:
        create_excel_calculator(m1, args.Kd, args.Ctotal, args.Nmax, args.excel)


if __name__ == "__main__":
    main()
