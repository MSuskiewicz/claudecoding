#!/usr/bin/env python3
"""
Read UniProt/TrEMBL IDs from an Excel file, fetch AlphaFold average pLDDT
scores, check for experimental PDB structures, and write results to Excel.
"""

import argparse
import json
import re
import sys
import time
import urllib.error
import urllib.request

from openpyxl import Workbook, load_workbook


UNIPROT_RE = re.compile(
    r"^[OPQ][0-9][A-Z0-9]{3}[0-9](?:-\d+)?$"    # UniProtKB accession
    r"|^[A-NR-Z][0-9]([A-Z][A-Z0-9]{2}[0-9]){1,2}(?:-\d+)?$"  # longer accessions
)


def is_uniprot_id(value):
    """Check if a string looks like a UniProt/TrEMBL accession."""
    return bool(UNIPROT_RE.match(value.strip()))


def strip_isoform(accession):
    """Remove isoform suffix (e.g. P12345-2 -> P12345)."""
    return re.sub(r"-\d+$", "", accession.strip())


def fetch_json(url, retries=2):
    """Fetch JSON from a URL with basic retry logic."""
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(url, timeout=30) as resp:
                return json.loads(resp.read())
        except (urllib.error.URLError, urllib.error.HTTPError, OSError, ValueError):
            if attempt < retries:
                time.sleep(1)
    return None


def get_protein_info(uniprot_id):
    """Fetch protein name and sequence length from UniProt.
    Returns (name, length_in_aa)."""
    data = fetch_json(f"https://rest.uniprot.org/uniprotkb/{uniprot_id}.json")
    if not data:
        return "N/A", "N/A"

    # Protein name
    desc = data.get("proteinDescription", {})
    rec = desc.get("recommendedName")
    if rec:
        name = rec.get("fullName", {}).get("value", "N/A")
    else:
        sub = desc.get("submissionNames", [{}])
        name = sub[0].get("fullName", {}).get("value", "N/A") if sub else "N/A"

    # Sequence length
    length = data.get("sequence", {}).get("length", "N/A")

    return name, length


def get_alphafold_plddt(uniprot_id):
    """Download the AlphaFold PDB file and compute average pLDDT from B-factors of CA atoms."""
    meta = fetch_json(f"https://alphafold.ebi.ac.uk/api/prediction/{uniprot_id}")
    if not meta:
        return None, None
    entry = meta[0] if isinstance(meta, list) else meta
    pdb_url = entry.get("pdbUrl")
    if not pdb_url:
        return None, None

    # Download the PDB text
    try:
        with urllib.request.urlopen(pdb_url, timeout=60) as resp:
            pdb_text = resp.read().decode("utf-8", errors="replace")
    except (urllib.error.URLError, urllib.error.HTTPError, OSError):
        return entry.get("entryId"), None

    # Parse CA atoms and extract pLDDT from B-factor column (cols 61-66)
    plddts = []
    for line in pdb_text.splitlines():
        if line.startswith("ATOM") and line[12:16].strip() == "CA":
            try:
                plddts.append(float(line[60:66]))
            except ValueError:
                continue

    if not plddts:
        return entry.get("entryId"), None
    return entry.get("entryId"), sum(plddts) / len(plddts)


def check_pdb_exists(uniprot_id):
    """Check RCSB PDB for experimental structures mapped to a UniProt accession.
    Returns a list of PDB IDs or an empty list."""
    url = "https://search.rcsb.org/rcsbsearch/v2/query"
    payload = {
        "query": {
            "type": "terminal",
            "service": "text",
            "parameters": {
                "attribute": "rcsb_polymer_entity_container_identifiers.reference_sequence_identifiers.database_accession",
                "operator": "exact_match",
                "value": uniprot_id,
            },
        },
        "return_type": "entry",
        "request_options": {"paginate": {"start": 0, "rows": 50}},
    }
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode(),
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
    except (urllib.error.URLError, urllib.error.HTTPError, OSError):
        return []
    return [hit["identifier"] for hit in data.get("result_set", [])]


def main():
    parser = argparse.ArgumentParser(
        description="For each UniProt ID in an Excel file, fetch AlphaFold average "
        "pLDDT and check for experimental PDB structures."
    )
    parser.add_argument("input", help="Input Excel file (.xlsx) with UniProt IDs in column A")
    parser.add_argument(
        "-o", "--output", default="folded_proteins_results.xlsx",
        help="Output Excel file (default: folded_proteins_results.xlsx)",
    )
    parser.add_argument(
        "--sheet", default=None,
        help="Sheet name to read (default: active sheet)",
    )
    parser.add_argument(
        "--skip-header", action="store_true",
        help="Skip the first row (header row)",
    )
    args = parser.parse_args()

    wb_in = load_workbook(args.input, read_only=True)
    ws_in = wb_in[args.sheet] if args.sheet else wb_in.active

    # Collect UniProt IDs from column A, skipping non-accession rows
    all_values = []
    for row in ws_in.iter_rows(min_col=1, max_col=1, values_only=True):
        val = row[0]
        if val is not None:
            all_values.append(str(val).strip())
    wb_in.close()

    if args.skip_header and all_values:
        all_values = all_values[1:]

    raw_ids = []
    skipped = 0
    for val in all_values:
        if is_uniprot_id(val):
            raw_ids.append(val)
        else:
            skipped += 1

    if not raw_ids:
        print("No valid UniProt IDs found in the input file.", file=sys.stderr)
        sys.exit(1)

    print(f"Found {len(raw_ids)} valid UniProt IDs from {args.input}"
          f" ({skipped} non-ID rows skipped)")

    # Prepare output
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Results"
    ws_out.append([
        "Original ID",
        "UniProt Accession",
        "Protein Name",
        "Size (AA)",
        "AlphaFold Entry",
        "Avg pLDDT",
        "pLDDT Category",
        "Experimental PDB",
        "PDB IDs",
    ])

    for idx, raw_id in enumerate(raw_ids, 1):
        uniprot_id = strip_isoform(raw_id)
        print(f"[{idx}/{len(raw_ids)}] {raw_id} -> {uniprot_id} ... ", end="", flush=True)

        protein_name, protein_length = get_protein_info(uniprot_id)
        af_entry, avg_plddt = get_alphafold_plddt(uniprot_id)
        pdb_ids = check_pdb_exists(uniprot_id)

        # Categorise pLDDT
        if avg_plddt is not None:
            if avg_plddt > 90:
                category = "Very high"
            elif avg_plddt > 70:
                category = "Confident"
            elif avg_plddt > 50:
                category = "Low"
            else:
                category = "Very low"
            plddt_display = round(avg_plddt, 2)
        else:
            category = "N/A"
            plddt_display = "N/A"

        has_pdb = "Yes" if pdb_ids else "No"
        pdb_str = ", ".join(pdb_ids) if pdb_ids else ""

        ws_out.append([
            raw_id,
            uniprot_id,
            protein_name,
            protein_length,
            af_entry or "Not found",
            plddt_display,
            category,
            has_pdb,
            pdb_str,
        ])
        print(f"{protein_name} ({protein_length} AA)  pLDDT={plddt_display}  PDB={'Yes' if pdb_ids else 'No'}")

    wb_out.save(args.output)
    print(f"\nResults saved to {args.output}")


if __name__ == "__main__":
    main()
